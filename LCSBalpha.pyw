#!/usr/bin/env python
# -*- coding: utf-8 -*-

import sys
from PyQt4 import QtCore, QtGui
from PyQt4.QtCore import *
from PyQt4.QtGui import *
import io, os, time, openpyxl, fnmatch, csv, datetime
from datetime import date
from openpyxl import Workbook
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
'''--------------------------ПЕРЕМЕННЫЕ------------------------------'''
#DropSense
#volume =[]
savefiletext = ''
usavefiletext = ''
TCL_ALL_EVENTS = 0
filenameDS = ''
filelistDS = []
namelist = []
seqlist = []
filetxt = ''
txtlist = []
ufileresultXLS = ''
listitems=[]
#XLS
fileNameXLS = ''
fileNameCSV = ''
sheets = []
selectvalue = ''
#CSV
fileName = ''
days = 0
space = []
projlist2 =[]
finalnumlist = []
proclist = []
'''-------------------------ОБЩИЕ ФУНКЦИИ------------------------------'''
'''--------------------------STATISTICA---------------------------------'''
'''----------------------------EXCEL------------------------------------'''
def getFileXLS():
    global ufileNameXLS
    fileNameXLS = QStringList()
    fileNameXLS = QtGui.QFileDialog.getOpenFileName(caption = 'Open file', directory='c:\\', filter = " Excel sheet (*.xlsx *.xls)") 
    ufileNameXLS = str(fileNameXLS.toUtf8()).decode("utf-8")
    wb2 = load_workbook(ufileNameXLS)
    for pos in wb2.sheetnames:
        scrollArea.addItem(pos)

def selectlistItem(item):   
    global selectvalue
    selectvalue = str(item.text().toUtf8()).decode("utf-8")
    
def stat(filename):
    wb = openpyxl.load_workbook(filename = ufileNameXLS)
    sheet = wb[selectvalue]
    project_list = []
    project_list_2 = []
    countbase_list =[]
    i = 2 
    val_base = 0
    val_proj = ''
    while val_proj != None:
        val_proj = sheet.cell(row=i, column=9).value
        project_list.append(val_proj)
        val_base = sheet.cell(row=i, column=4).value
        countbase_list.append(val_base)
        i+=1
    del project_list[-1]
    del countbase_list[-1] 
    countbase_list_2 = [int(v) for v in countbase_list]
    for j in project_list:
           if j not in project_list_2:
               project_list_2.append(j)   
    j=0
    final_countbase_list = []
    while j<len(project_list_2):
        k=0
        i=0
        while i<len(project_list):
            if project_list_2[j] == project_list[i]:
                k = k + countbase_list_2[i]
            i+=1
        final_countbase_list.append(k)
        j+=1
    varsum = sum(final_countbase_list)
    # ЗАПИСЬ РЕЗУЛЬТАТОВ В ФАЙЛ
    varsum = sum(final_countbase_list)
    sheet.cell(row=1, column=12).value = 'Project'
    sheet.cell(row=1, column=13).value = 'Total Lenght'
    sheet.cell(row=1, column=14).value = '%'
    m=1
    while m < len(project_list_2)+1:
        sheet.cell(row=m+1, column=12).value = project_list_2[m-1]
        sheet.cell(row=m+1, column=13).value = final_countbase_list[m-1]
        sheet.cell(row=m+1, column=14).value = round(((float(final_countbase_list[m-1])/varsum)*100), 2)
        m=m+1
    wb.save(ufileNameXLS)
    msg=QMessageBox()
    msg.setIcon(QMessageBox.Information)
    msg.setText(u"Процесс успешно завершен")
    msg.setInformativeText(u"Откройте файл для просмотра результата")
    msg.setWindowTitle(u"Работа скрипта")
    msg.setDetailedText(u"Работа скрипта выполнена без ошибок")
    msg.setStandardButtons(QMessageBox.Ok | QMessageBox.Cancel)
    retval=msg.exec_()


'''-----------------------------CSV--------------------------------------'''
def getdirectory():
    global ufileName
    fileName = QStringList()
    fileName = QtGui.QFileDialog.getExistingDirectory(directory = u"/") 
    ufileName = str(fileName.toUtf8()).decode("utf-8")

def getweek():
    global space
    space = []
    today = datetime.date.today()
    global days #ЗАБРАТЬ ИЗ ИНТЕФЕЙСА
    days = int(spinBox.value())
    print days
    for day in range(days):
        space.append((str(today - datetime.timedelta(days=day))).replace("-", ""))

def ohibkifile():
    os.chdir(ufileName)
    path = os.getcwd()
    numlist = []
    projlist = []
    matches =[]
    ohibki = []
    ohibki2 = []
# ЗАПОЛНЕНИЕ МАССИВА ИМЕНАМИ .CSV-файлов
    for root, dirnames, filenames in os.walk(path):
        for filename in fnmatch.filter(filenames, '*.csv'):
            if(filename.find('F') <> 0 and filename.find('Re') <> 0 and filename.find('plate') < 0 and filename.find('flankers') <> 0):
                matches.append(os.path.join(root, filename))
#ОБРАБОТА CSV-ФАЙЛОВ                 
    for _match in matches:
        '''num = 0'''
        with open(_match, 'rb') as csvfile:
            spamreader = csv.DictReader(csvfile, delimiter=';', quotechar='|')
            for row in spamreader:
                '''if row['Concentration, ng/ul'] == '':
                    ohibki.append('PROJ_' + _match)'''
                try:
                    int(row['Length'])
                except ValueError:
                    ohibki.append('OLIG_ValueError' + _match)
                except KeyError:
                    ohibki.append('OLIG_KeyError' + _match)
                try:
                    row['Concentration, ng/ul']
                except ValueError:
                    ohibki.append('PROJ_ValueError' + _match)
                    break
                except KeyError:
                    ohibki.append('PROJ_KeyError' + _match)
                    break
                except UnicodeError:
                    ohibki.append('PROJ_KeyError' + _match)
                    break
                if row['Concentration, ng/ul'] == '':
                    ohibki.append('PROJ_NullObject' + _match)
#ВЫВОД В ТЕКСТОВЫЙ ФАЙЛ
    for j in ohibki:
        if j not in ohibki2:
            ohibki2.append(j)
    filetxt = open("C:\Temp\out.txt", "w+")
    for item in ohibki2:
        filetxt.write("%s\n" % item)
    filetxt.close()
    os.system("start "+"C:\Temp\out.txt")

def statweek():
    os.chdir(ufileName)
    path = os.getcwd()
    numlist = []
    projlist = []
    matches =[]
    num = 0
# ЗАПОЛНЕНИЕ МАССИВА ИМЕНАМИ .CSV-файлов
    for item in space:
        for root, dirnames, filenames in os.walk(path):
            for filename in fnmatch.filter(filenames, '*.csv'):
                if(filename.find(item) >= 0 and filename.find('F') <> 0 and filename.find('plate') < 0 and filename.find('Re') <> 0 and filename.find('flankers') <> 0):
                    matches.append(os.path.join(root, filename))
                    
#ОБРАБОТА CSV-ФАЙЛОВ                 
    for _match in matches:
        num = 0
        with open(_match, 'rb') as csvfile:
            spamreader = csv.DictReader(csvfile, delimiter=';', quotechar='|')
            for row in spamreader:
                num = num + int(row['Length'])
            numlist.append(num)
            projlist.append(row['Concentration, ng/ul'])
# ОБРАБОТКА И ФОРМИРОВАНИЕ КОНЕЧНОГО МАССИВА
    j = 0
    global projlist2
    global finalnumlist
    projlist2 = []
    for j in projlist:
        if j not in projlist2:
            projlist2.append(j)
    j=0
    finalnumlist = []
    while j<len(projlist2):
        k=0
        i=0
        while i<len(projlist):
            if projlist2[j] == projlist[i]:
                k = k + numlist[i]
            i+=1
        finalnumlist.append(k)
        j+=1
# ЗАПИСЬ В EXCEL ФАЙЛ
    varsum = sum(finalnumlist)
    wb = Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1).value = 'Project'
    sheet.cell(row=1, column=2).value = 'Total Lenght'
    sheet.cell(row=1, column=3).value = '%'
    m=1
    while m < len(projlist2)+1:
        sheet.cell(row=m+1, column=1).value = projlist2[m-1]
        sheet.cell(row=m+1, column=2).value = finalnumlist[m-1]
        sheet.cell(row=m+1, column=3).value = round(((float(finalnumlist[m-1])/varsum)*100), 2)
        m=m+1
    wb.save("C:\Temp\sample.xlsx")
    os.system("start "+"C:\Temp\sample.xlsx")

def statdir():
    os.chdir(ufileName)
    path = os.getcwd()
    numlist = []
    projlist = []
    matches =[]
    num = 0
# ЗАПОЛНЕНИЕ МАССИВА ИМЕНАМИ .CSV-файлов
    for root, dirnames, filenames in os.walk(path):
        for filename in fnmatch.filter(filenames, '*.csv'):
            if(filename.find('F') <> 0 and filename.find('Re') <> 0 and filename.find('plate') < 0 and filename.find('flankers') <> 0):
                matches.append(os.path.join(root, filename))
#ОБРАБОТА CSV-ФАЙЛОВ
    for _match in matches:
        num = 0
        with open(_match, 'rb') as csvfile:
            spamreader = csv.DictReader(csvfile, delimiter=';', quotechar='|')
            for row in spamreader:
                num = num + int(row['Length'])
            numlist.append(num)
            projlist.append(row['Concentration, ng/ul'])
# ОБРАБОТКА И ФОРМИРОВАНИЕ КОНЕЧНОГО МАССИВА
    j = 0
    global projlist2
    global finalnumlist
    projlist2 = []
    for j in projlist:
        if j not in projlist2:
            projlist2.append(j)
    j=0
    finalnumlist = []
    while j<len(projlist2):
        k=0
        i=0
        while i<len(projlist):
            if projlist2[j] == projlist[i]:
                k = k + numlist[i]
            i+=1
        finalnumlist.append(k)
        j+=1
# ЗАПИСЬ В EXCEL ФАЙЛ
    varsum = sum(finalnumlist)
    wb = Workbook()
    sheet = wb.active
    sheet.cell(row=1, column=1).value = 'Project'
    sheet.cell(row=1, column=2).value = 'Total Lenght'
    sheet.cell(row=1, column=3).value = '%'
    m=1
    while m < len(projlist2)+1:
        sheet.cell(row=m+1, column=1).value = projlist2[m-1]
        sheet.cell(row=m+1, column=2).value = finalnumlist[m-1]
        sheet.cell(row=m+1, column=3).value = round(((float(finalnumlist[m-1])/varsum)*100), 2)
        m=m+1
    wb.save("C:\Temp\sample.xlsx")
    os.system("start "+"C:\Temp\sample.xlsx")

def grafoniy():
    N = len(projlist2)
    value = finalnumlist
    verh = int(max(finalnumlist) + 0.15*max(finalnumlist))
    gradient = int(max(finalnumlist)/10)
    proclist = finalnumlist
    ind = np.arange(N)
    width = 0.75
    plt.figure(1)
    p1 = plt.bar(ind, value, width)
    plt.ylabel('Count of base')
    plt.xlabel('Project')
    plt.title('Statistica proektov')
    plt.xticks(ind, projlist2)
    plt.yticks(np.arange(0, verh, gradient))
    plt.figure(2)
    plt.pie(proclist, labels = projlist2, autopct='%1.1f%%')
    plt.show()

'''------------------------------DROPSENSE--------------------------------'''

def getresultFile():
    global ufileresultXLS
    fileresultXLS = QtGui.QFileDialog.getOpenFileName(caption = 'Open file', directory='c:\\', filter = " Excel sheet (*.csv)")
    ufileresultXLS = str(fileresultXLS.toUtf8()).decode("utf-8")

def formDSworklist():
    statusBar.showMessage(u"Сохранение варклиста" ,4000)
    worklistPath = QtGui.QFileDialog.getSaveFileName(filter = "Excel sheet (*.xlsx)")
    workbookDS.save(str(worklistPath.toUtf8()).decode("utf-8"))
    statusBar.showMessage(u"Ворклист сохранен",4000)
    
def addlistItem():
    listArea.addItem(box.selectedFiles()[0])
    
def dellistItem():
    item = listArea.takeItem(listArea.currentRow())
    item = None
    
def saveItem():
    global listitems
    global workbookDS
    listitems = []
    listitem = ''
    for index in range(listArea.count()):
        listitem = str(listArea.item(index).text().toUtf8()).decode("utf-8")
        listitems.append(listitem)
    namelist = []
    seqlist = []
    for _match in listitems:
        with open(_match, 'rb') as csvfile:
            spamreader = csv.DictReader(csvfile, delimiter=';', quotechar='|')
            for row in spamreader:
                namelist.append(row['Name'])
                seqlist.append(row['Sequence'])
    workbookDS = openpyxl.load_workbook(filename = 'input2.xlsx')
    sheet = workbookDS.active
    #ВСЁ ПОПРАВИТЬ
    m=comboBox.currentIndex()+1
    o=0
    while o+1 < len(seqlist)+1:
        sheet.cell(row=m+1, column=5).value = namelist[0+o]
        sheet.cell(row=m+1, column=8).value = seqlist[0+o]
        sheet.cell(row=m+1, column=9).value = 'DNA'
        m=m+1
        o=o+1 #   :-o
    msg=QMessageBox()
    msg.setIcon(QMessageBox.Information)
    msg.setText(u"Готово. Закройте окно и \n сохраните worklist")
    msg.setStandardButtons(QMessageBox.Ok)
    retval=msg.exec_()

'''def comboIndex():
    print comboBox.currentIndex()'''
    
def getCSVDS():
    global listArea
    global box
    box = QtGui.QFileDialog(directory = " W:\/", filter = "CSV Files (*.csv)")
    listArea = QtGui.QListWidget(box)
    bCancel = QtGui.QPushButton(box)
    bCancel.setEnabled(False)
    bCancel.isFlat()
    bCancel.setStyleSheet("background-color: #f0f0f0")
    bCancel2 = QtGui.QPushButton(box)
    bCancel2.setEnabled(False)
    bCancel2.isFlat()
    bCancel2.setStyleSheet("background-color: #f0f0f0")
    bSave = QtGui.QPushButton(box)
    bSave.setText(u"Сохранить выбранные CSV")
    bSave.clicked.connect(saveItem)
    bAdd = QtGui.QPushButton(box)
    bAdd.setText(u"Добавить")
    bAdd.clicked.connect(addlistItem)
    bDel = QtGui.QPushButton(box)
    bDel.setText(u"Удалить")
    bDel.clicked.connect(dellistItem)
    layout = box.layout()
    layout.addWidget(listArea, 0, 3, 2, 2)
    layout.addWidget(bAdd, 2, 3, 1, 2)
    layout.addWidget(bDel, 3, 3, 1, 2)
    layout.addWidget(bSave, 4, 3, 1, 2)
    layout.addWidget(bCancel, 2, 2, 1, 1)
    layout.addWidget(bCancel2, 2, 2, 3, 1)
    retval = box.exec_()

def putworklistDS():
    global resultstr
    resultstr = ''
    global result
    result = []
    global volume
    volume =[]
    '''wb2 = openpyxl.load_workbook(filename = ufileresultXLS)
    sheet = wb2.active
    i=2
    while resultstr != None:
        resultstr = sheet.cell(row=i, column=1).value
        result.append(resultstr)
        i+=1
    del result[-1], result[-1]'''
    res = []
    with open(ufileresultXLS, 'rb') as csvfile:
        spamreader = csv.reader(csvfile, delimiter=';', quotechar='|')
        for row in spamreader:
            res.append(row)
    for i in range(len(res)):
        result.append(res[i][0])
        result[i] = result[i].replace(".",",")
    del result[0], result[-1]
    source = []
    z=1
    global r
    r=0
    for x in range(len(listitems)):
        with open(listitems[x]) as File:
            reader = csv.reader(File, delimiter=';', quotechar='|', quoting=csv.QUOTE_MINIMAL)
            source = []
            for row in reader:
                source.append(row)
            for z in range(len(source)-1):
                source[z+1][9] = str(int((float(result[r].replace(',','.'))*int(spinBox_2.value()))/(int(spinBox_3.value())) - int(spinBox_2.value()))+int(spinBox_2.value()))
                source[z+1][10] = str(round((float(result[r].replace(',','.'))*int(spinBox_2.value()))/1000, 2)).replace('.',',')
                source[z+1][12] = str(result[r])
                volume.append(str(int((float(result[r].replace(',','.'))*int(spinBox_2.value()))/(int(spinBox_3.value())) - int(spinBox_2.value()))))
                r=r+1
            source[0][9] = 'Volume'
            source[0][10] = 'nmol'
            source[0][12] = 'Concentration, pmol/ul'
        with open(listitems[x], 'wb') as csvfile:
            spamwriter = csv.writer(csvfile, delimiter=';') 
            spamwriter.writerows(source)
    wbk = Workbook()
    sheet = wbk.active
    t=1
    while t < len(volume)+1:
        sheet.cell(row=t+1, column=1).value = volume[t-1]
        t=t+1
    sheet.cell(row=1, column=1).value = u'Объем воды'
    wbk.save("C:\Temp\K_" + str(spinBox_3.value()) + "_pmol.xlsx")
    os.system("start " + "C:\Temp\K_" + str(spinBox_3.value()) + "_pmol.xlsx")
    msg=QMessageBox()
    msg.setIcon(QMessageBox.Information)
    msg.setText(u"Готово")
    msg.setStandardButtons(QMessageBox.Ok)
    retval=msg.exec_()

def getTXTDSFile():
    global filetxt
    global txtlist
    txtlist = []
    filetxt = QtGui.QFileDialog.getOpenFileName(caption = 'Open file', directory='c:\\', filter = " Text file (*.txt)")
    ufiletxt = str(filetxt.toUtf8()).decode("utf-8")
    txt = io.open(ufiletxt, mode='r').read().split()
    i=1
    while i<len(txt):
        txtlist.append(txt[i])
        i=i+3
        
def putTXTDSFile():
    global savefiletext
    global usavefiletext
    savefiletext = QtGui.QFileDialog.getSaveFileName(filter = "Text file (*.txt)")
    usavefiletext = str(savefiletext.toUtf8()).decode("utf-8")
    filetext = open(usavefiletext, "w+")
    for item in txtlist:
        filetext.write("%s\n" % str(item))
    filetext.close()

def putRe():
    sourceRe = []
    sourceRe2 = []
    z=1
    global r
    r=0
    for x in range(len(listitems)):
        with open(listitems[x]) as File:
            reader = csv.reader(File, delimiter=';', quotechar='|', quoting=csv.QUOTE_MINIMAL)
            sourceRe = []
            volume = []
            for row in reader:
                sourceRe.append(row)
            for z in range(len(sourceRe)-1):
                if ((50*0.98**(int(sourceRe[z+1][5])-1))>float(sourceRe[z+1][10].replace(',','.'))):
                    sourceRe2.append(sourceRe[z+1][:])
                    r=r+1
        if(len(sourceRe2)>0):
            sourceRe2.insert(0, ['Customer','Plate','WellPosition','Name','Sequence','Length','Scale','MW','Ext.Coef','Volume','nmol','Concentration, ng/ul','Concentration, pmol/ul',''])
            with open(listitems[x][0:listitems[x].rfind('/')+1] + 'Re_' + listitems[x][listitems[x].rfind('/')+1:], 'wb') as csvfile:  
                spamwriter = csv.writer(csvfile, delimiter=';') 
                spamwriter.writerows(sourceRe2)
    if r>0:
        msg=QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText(u"Re файлы сформированы")   
        msg.setStandardButtons(QMessageBox.Ok)
        retval=msg.exec_()
    else:
        msg=QMessageBox()
        msg.setIcon(QMessageBox.Information)
        msg.setText(u"Вылетов по концентрации не обнаружено")   
        msg.setStandardButtons(QMessageBox.Ok)
        retval=msg.exec_()

def putF():
    sourceF = []
    sourceF2 = []
    z=1
    global r
    r=0
    for x in range(len(listitems)):
        with open(listitems[x]) as File:
            reader = csv.reader(File, delimiter=';', quotechar='|', quoting=csv.QUOTE_MINIMAL)
            sourceF = []
            volume = []
            for row in reader:
                sourceF.append(row)
            for z in range(len(sourceF)-1):
                sourceF[z+1][12]='100'
                del sourceF[z+1][1], sourceF[z+1][2]
        sourceF[0]=['Customer', 'Name','Sequence','Length','Scale','MW','Ext.Coef','Volume','nmol','Concentration, ng/ul','Concentration, pmol/ul','']
        with open(listitems[x][0:listitems[x].rfind('/')+1] + 'F_' + listitems[x][listitems[x].rfind('/')+1:], 'wb') as csvfile:
            spamwriter = csv.writer(csvfile, delimiter=';') 
            spamwriter.writerows(sourceF)
    msg=QMessageBox()
    msg.setIcon(QMessageBox.Information)
    msg.setText(u"F файлы сформированы")   
    msg.setStandardButtons(QMessageBox.Ok)
    retval=msg.exec_()
'''---------------------------УТИЛИТЫ----------------------------'''
def loadWL():
    global filetxtWL
    global txtlistWL
    txtlistWL = []
    filetxtWL = QtGui.QFileDialog.getOpenFileName(caption = 'Open file', directory='c:\\', filter = " Text file (*.txt)")
    ufiletxtWL = str(filetxtWL.toUtf8()).decode("utf-8")
    txt = io.open(ufiletxtWL, mode='r').read().split()
    i=1
    while i<len(txt):
        txtlistWL.append(txt[i])
        i=i+3
    #print txtlistWL

def rawforSynthesis():
    Mon_dA = 0
    Mon_dC = 0
    Mon_dG = 0
    Mon_dT = 0
    Mon_All = 0
    for oligo in range(len(txtlistWL)):
        for mon in range(len(txtlistWL[oligo])):
            Mon_All+=1
            if str(txtlistWL[oligo][mon]) == 'A':
                Mon_dA+=1
            elif str(txtlistWL[oligo][mon]) == 'C':
                Mon_dC+=1
            elif str(txtlistWL[oligo][mon]) == 'G':
                Mon_dG+=1
            elif str(txtlistWL[oligo][mon]) == 'T':
                Mon_dT+=1

    print Mon_All, Mon_dA, Mon_dC, Mon_dG, Mon_dT
             

'''-----------------------------UI-------------------------------'''
app = QtGui.QApplication(sys.argv)
LCSBSuite = QtGui.QMainWindow()
LCSBSuite.setObjectName(("LCSBSuite"))
LCSBSuite.setWindowModality(QtCore.Qt.NonModal)
LCSBSuite.resize(776, 541)
LCSBSuite.setMinimumSize(QtCore.QSize(776, 541))
LCSBSuite.setMaximumSize(QtCore.QSize(776, 541))
font = QtGui.QFont()
font.setPointSize(10)
LCSBSuite.setFont(font)
LCSBSuite.setDocumentMode(False)
LCSBSuite.setTabShape(QtGui.QTabWidget.Rounded)
centralwidget = QtGui.QWidget(LCSBSuite)
centralwidget.setObjectName(("centralwidget"))
horizontalLayout = QtGui.QHBoxLayout(centralwidget)
horizontalLayout.setObjectName(("horizontalLayout"))
verticalLayout = QtGui.QVBoxLayout()
verticalLayout.setObjectName(("verticalLayout"))
tabWidget = QtGui.QTabWidget(centralwidget)
tabWidget.setFont(font)
tabWidget.setContextMenuPolicy(QtCore.Qt.DefaultContextMenu)
tabWidget.setLayoutDirection(QtCore.Qt.LeftToRight)
tabWidget.setObjectName(("tabWidget"))
tab = QtGui.QWidget()
tab.setObjectName(("tab"))
verticalLayout_7 = QtGui.QVBoxLayout(tab)
verticalLayout_7.setObjectName(("verticalLayout_7"))
CSVgroupBox = QtGui.QGroupBox(tab)
CSVgroupBox.setObjectName(("CSVgroupBox"))
horizontalLayout_8 = QtGui.QHBoxLayout(CSVgroupBox)
horizontalLayout_8.setObjectName(("horizontalLayout_8"))
horizontalLayout_7 = QtGui.QHBoxLayout()
horizontalLayout_7.setObjectName(("horizontalLayout_7"))
verticalLayout_3 = QtGui.QVBoxLayout()
verticalLayout_3.setObjectName(("verticalLayout_3"))
selcsvButton = QtGui.QPushButton(CSVgroupBox)
selcsvButton.setFont(font)
selcsvButton.setObjectName(("selcsvButton"))
selcsvButton.clicked.connect(getCSVDS)
verticalLayout_3.addWidget(selcsvButton)
wlDSButton = QtGui.QPushButton(CSVgroupBox)
wlDSButton.setFont(font)
wlDSButton.setObjectName(("wlDSButton"))
wlDSButton.clicked.connect(formDSworklist)
verticalLayout_3.addWidget(wlDSButton)
horizontalLayout_4 = QtGui.QHBoxLayout()
horizontalLayout_4.setObjectName(("horizontalLayout_4"))
startcell = QtGui.QLabel(CSVgroupBox)
startcell.setObjectName(("startcell"))
horizontalLayout_4.addWidget(startcell)
comboBox = QtGui.QComboBox(CSVgroupBox)
for j in range(1, 13):
    for i in ["A","B","C","D","E","F","G","H"]:
        comboBox.addItem(i + str(j))
comboBox.setModelColumn(0)
comboBox.setObjectName(("comboBox"))
horizontalLayout_4.addWidget(comboBox)
verticalLayout_3.addLayout(horizontalLayout_4)
verticalLayout_2 = QtGui.QVBoxLayout()
verticalLayout_2.setObjectName(("verticalLayout_2"))
seqRButton = QtGui.QRadioButton(CSVgroupBox)
seqRButton.setChecked(True)
seqRButton.setObjectName(("seqRButton"))
verticalLayout_2.addWidget(seqRButton)
mwRButton = QtGui.QRadioButton(CSVgroupBox)
mwRButton.setObjectName(("mwRButton"))
verticalLayout_2.addWidget(mwRButton)
spacerItem = QtGui.QSpacerItem(20, 40, QtGui.QSizePolicy.Minimum, QtGui.QSizePolicy.Expanding)
verticalLayout_2.addItem(spacerItem)
verticalLayout_3.addLayout(verticalLayout_2)
horizontalLayout_7.addLayout(verticalLayout_3)
verticalLayout_6 = QtGui.QVBoxLayout()
verticalLayout_6.setObjectName(("verticalLayout_6"))
resDSButton = QtGui.QPushButton(CSVgroupBox)
resDSButton.clicked.connect(getresultFile)
resDSButton.setObjectName(("resDSButton"))
verticalLayout_6.addWidget(resDSButton)
dataButton = QtGui.QPushButton(CSVgroupBox)
dataButton.clicked.connect(putworklistDS)
dataButton.setObjectName(("dataButton"))
verticalLayout_6.addWidget(dataButton)
horizontalLayout_2 = QtGui.QHBoxLayout()
horizontalLayout_2.setObjectName(("horizontalLayout_2"))
label_2 = QtGui.QLabel(CSVgroupBox)
label_2.setObjectName(("label_2"))
horizontalLayout_2.addWidget(label_2)
spinBox_2 = QtGui.QSpinBox(CSVgroupBox)
spinBox_2.setMaximum(1000)
spinBox_2.setProperty("value", 98)
spinBox_2.setObjectName(("spinBox_2"))
horizontalLayout_2.addWidget(spinBox_2)
label_5 = QtGui.QLabel(CSVgroupBox)
label_5.setObjectName(("label_5"))
horizontalLayout_2.addWidget(label_5)
verticalLayout_6.addLayout(horizontalLayout_2)
horizontalLayout_3 = QtGui.QHBoxLayout()
horizontalLayout_3.setObjectName(("horizontalLayout_3"))
label_3 = QtGui.QLabel(CSVgroupBox)
label_3.setObjectName(("label_3"))
horizontalLayout_3.addWidget(label_3)
spinBox_3 = QtGui.QSpinBox(CSVgroupBox)
spinBox_3.setMinimum(50)
spinBox_3.setMaximum(100)
spinBox_3.setSingleStep(50)
spinBox_3.setProperty("value", 100)
spinBox_3.setObjectName(("spinBox_3"))
horizontalLayout_3.addWidget(spinBox_3)
label_4 = QtGui.QLabel(CSVgroupBox)
label_4.setObjectName(("label_4"))
horizontalLayout_3.addWidget(label_4)
verticalLayout_6.addLayout(horizontalLayout_3)
Re_Button = QtGui.QPushButton(CSVgroupBox)
Re_Button.clicked.connect(putRe)
Re_Button.setObjectName(("Re_Button"))
verticalLayout_6.addWidget(Re_Button)
F_Button = QtGui.QPushButton(CSVgroupBox)
F_Button.clicked.connect(putF)
F_Button.setObjectName("F_Button")
verticalLayout_6.addWidget(F_Button)
horizontalLayout_7.addLayout(verticalLayout_6)
horizontalLayout_8.addLayout(horizontalLayout_7)
verticalLayout_7.addWidget(CSVgroupBox)
txtgroupBox = QtGui.QGroupBox(tab)
txtgroupBox.setObjectName(("txtgroupBox"))
horizontalLayout_10 = QtGui.QHBoxLayout(txtgroupBox)
horizontalLayout_10.setObjectName(("horizontalLayout_10"))
horizontalLayout_9 = QtGui.QHBoxLayout()
horizontalLayout_9.setObjectName(("horizontalLayout_9"))
txtopenButton = QtGui.QPushButton(txtgroupBox)
txtopenButton.clicked.connect(getTXTDSFile)
txtopenButton.setObjectName(("txtopenButton"))
horizontalLayout_9.addWidget(txtopenButton)
txtsaveButton = QtGui.QPushButton(txtgroupBox)
txtsaveButton.clicked.connect(putTXTDSFile)
txtsaveButton.setObjectName(("txtsaveButton"))
horizontalLayout_9.addWidget(txtsaveButton)
horizontalLayout_10.addLayout(horizontalLayout_9)
verticalLayout_7.addWidget(txtgroupBox)
tabWidget.addTab(tab, (""))
'''--------------------------ВТОРАЯ ВКЛАДКА----------------------------'''
tab_2 = QtGui.QWidget()
tab_2.setObjectName(("tab_2"))
gridLayout_2 = QtGui.QGridLayout(tab_2)
gridLayout_2.setObjectName(("gridLayout_2"))
graphBox = QtGui.QGroupBox(tab_2)
graphBox.setObjectName(("graphBox"))
gridLayout_3 = QtGui.QGridLayout(graphBox)
gridLayout_3.setObjectName(("gridLayout_3"))
spacerItem3 = QtGui.QSpacerItem(20, 40, QtGui.QSizePolicy.Minimum, QtGui.QSizePolicy.Ignored)
gridLayout_3.addItem(spacerItem3, 2, 1, 1, 1)
showgrafButton = QtGui.QPushButton(graphBox)
showgrafButton.clicked.connect(grafoniy)
showgrafButton.setObjectName(("showgrafButton"))
gridLayout_3.addWidget(showgrafButton, 1, 1, 1, 1)
spacerItem4 = QtGui.QSpacerItem(20, 40, QtGui.QSizePolicy.Minimum, QtGui.QSizePolicy.Ignored)
gridLayout_3.addItem(spacerItem4, 0, 1, 1, 1)
spacerItem5 = QtGui.QSpacerItem(40, 20, QtGui.QSizePolicy.Expanding, QtGui.QSizePolicy.Minimum)
gridLayout_3.addItem(spacerItem5, 1, 0, 1, 1)
spacerItem6 = QtGui.QSpacerItem(40, 20, QtGui.QSizePolicy.Expanding, QtGui.QSizePolicy.Minimum)
gridLayout_3.addItem(spacerItem6, 1, 2, 1, 1)
gridLayout_2.addWidget(graphBox, 1, 1, 1, 1)
csvBox = QtGui.QGroupBox(tab_2)
csvBox.setAcceptDrops(False)
csvBox.setAutoFillBackground(False)
csvBox.setInputMethodHints(QtCore.Qt.ImhNone)
csvBox.setFlat(False)
csvBox.setCheckable(False)
csvBox.setObjectName(("csvBox"))
verticalLayout_4 = QtGui.QVBoxLayout(csvBox)
verticalLayout_4.setObjectName(("verticalLayout_4"))
#КНОПКА ВЫБОРА ДИРЕКТОРИИ
dirButton = QtGui.QPushButton(csvBox)
dirButton.clicked.connect(getdirectory)
dirButton.setObjectName(("dirButton"))
verticalLayout_4.addWidget(dirButton)
#КНОПКА ПРОВЕРКИ НА ОЩИБКИ
misButton = QtGui.QPushButton(csvBox)
misButton.clicked.connect(ohibkifile)
misButton.setObjectName(("misButton"))
verticalLayout_4.addWidget(misButton)
horizontalLayout_5 = QtGui.QHBoxLayout()
horizontalLayout_5.setObjectName(("horizontalLayout_5"))
label = QtGui.QLabel(csvBox)
label.setObjectName(("label"))
horizontalLayout_5.addWidget(label)
spinBox = QtGui.QSpinBox(csvBox)
spinBox.setMinimum(1)
spinBox.setProperty("value", 7)
spinBox.setObjectName(("spinBox"))
horizontalLayout_5.addWidget(spinBox)
spacerItem7 = QtGui.QSpacerItem(20, 20, QtGui.QSizePolicy.Expanding, QtGui.QSizePolicy.Minimum)
horizontalLayout_5.addItem(spacerItem7)
verticalLayout_4.addLayout(horizontalLayout_5)
intervalButton = QtGui.QPushButton(csvBox)
intervalButton.clicked.connect(getweek)
intervalButton.setObjectName(("intervalButton"))
verticalLayout_4.addWidget(intervalButton)
intstatButton = QtGui.QPushButton(csvBox)
intstatButton.clicked.connect(statweek)
intstatButton.setObjectName(("intstatButton"))
verticalLayout_4.addWidget(intstatButton)
dirstatButton = QtGui.QPushButton(csvBox)
#КНОПКА СТАТИСТИКИ ДИРЕКТОРИИ
dirstatButton.setObjectName(("dirstatButton"))
dirstatButton.clicked.connect(statdir)
verticalLayout_4.addWidget(dirstatButton)
gridLayout_2.addWidget(csvBox, 0, 0, 2, 1)
#ЭКСЭЛЬ СТАТИСТИКА
excelBox = QtGui.QGroupBox(tab_2)
excelBox.setObjectName(("excelBox"))
verticalLayout_5 = QtGui.QVBoxLayout(excelBox)
verticalLayout_5.setObjectName(("verticalLayout_5"))
excelopenButton = QtGui.QPushButton(excelBox)
excelopenButton.clicked.connect(getFileXLS)
excelopenButton.setObjectName(("excelopenButton"))
verticalLayout_5.addWidget(excelopenButton)
scrollArea = QtGui.QListWidget(excelBox) 
scrollArea.itemClicked.connect(selectlistItem)
verticalLayout_5.addWidget(scrollArea)
excelButton2 = QtGui.QPushButton(excelBox)
excelButton2.clicked.connect(stat)
excelButton2.setObjectName(("excelButton2"))
verticalLayout_5.addWidget(excelButton2)
gridLayout_2.addWidget(excelBox, 0, 1, 1, 1)
tabWidget.addTab(tab_2,(""))
'''-----------------------------ТРЕТЬЯ ВКЛАДКА-------------------------------'''
tab_3 = QtGui.QWidget()
tab_3.setObjectName("tab_3")
verticalLayout_8 = QtGui.QVBoxLayout(tab_3)
verticalLayout_8.setObjectName("verticalLayout_8")
groupBox = QtGui.QGroupBox(tab_3)
groupBox.setObjectName("groupBox")
horizontalLayout_15 = QtGui.QHBoxLayout(groupBox)
horizontalLayout_15.setObjectName("horizontalLayout_15")
horizontalLayout_14 = QtGui.QHBoxLayout()
horizontalLayout_14.setObjectName("horizontalLayout_14")
loadWLButton = QtGui.QPushButton(groupBox)
loadWLButton.clicked.connect(loadWL)
loadWLButton.setObjectName("loadWLButton")
horizontalLayout_14.addWidget(loadWLButton)
horizontalLayout_13 = QtGui.QHBoxLayout()
horizontalLayout_13.setObjectName("horizontalLayout_13")
horizontalLayout_11 = QtGui.QHBoxLayout()
horizontalLayout_11.setObjectName("horizontalLayout_11")
label_6 = QtGui.QLabel(groupBox)
label_6.setObjectName("label_6")
horizontalLayout_11.addWidget(label_6)
comboBox_2 = QtGui.QComboBox(groupBox)
for cB2 in ['2', '5', '10']:
    comboBox_2.addItem(cB2)
comboBox_2.setCurrentIndex(1)
comboBox_2.setObjectName("comboBox_2")
horizontalLayout_11.addWidget(comboBox_2)
horizontalLayout_13.addLayout(horizontalLayout_11)
horizontalLayout_12 = QtGui.QHBoxLayout()
horizontalLayout_12.setObjectName("horizontalLayout_12")
label_7 = QtGui.QLabel(groupBox)
label_7.setObjectName("label_7")
horizontalLayout_12.addWidget(label_7)
comboBox_3 = QtGui.QComboBox(groupBox)
for cB3 in ['1', '2', '3', '4']:
    comboBox_3.addItem(cB3)
comboBox_3.setCurrentIndex(2)
comboBox_3.setObjectName("comboBox_3")
horizontalLayout_12.addWidget(comboBox_3)
horizontalLayout_13.addLayout(horizontalLayout_12)
horizontalLayout_14.addLayout(horizontalLayout_13)
pushButton_2 = QtGui.QPushButton(groupBox)
pushButton_2.clicked.connect(rawforSynthesis)
pushButton_2.setObjectName("pushButton_2")
horizontalLayout_14.addWidget(pushButton_2)
horizontalLayout_15.addLayout(horizontalLayout_14)
verticalLayout_8.addWidget(groupBox)
groupBox_3 = QtGui.QGroupBox(tab_3)
groupBox_3.setObjectName("groupBox_3")
horizontalLayout_18 = QtGui.QHBoxLayout(groupBox_3)
horizontalLayout_18.setObjectName("horizontalLayout_18")
horizontalLayout_16 = QtGui.QHBoxLayout()
horizontalLayout_16.setObjectName("horizontalLayout_16")
label_8 = QtGui.QLabel(groupBox_3)
label_8.setObjectName("label_8")
horizontalLayout_16.addWidget(label_8)
comboBox_4 = QtGui.QComboBox(groupBox_3)
for cB4 in [u'Синтез олигонуклеотидов', u'Синтез модифицированых олигонуклеотидов', u'Синтез РНК', u'Синтез ампликона']:
    comboBox_4.addItem(cB4)
comboBox_4.setMinimumSize(QtCore.QSize(380, 0))
comboBox_4.setObjectName("comboBox_4")
horizontalLayout_16.addWidget(comboBox_4)
pushButton_3 = QtGui.QPushButton(groupBox_3)
pushButton_3.setObjectName("pushButton_3")
horizontalLayout_16.addWidget(pushButton_3)
horizontalLayout_18.addLayout(horizontalLayout_16)
verticalLayout_8.addWidget(groupBox_3)
groupBox_2 = QtGui.QGroupBox(tab_3)
groupBox_2.setObjectName("groupBox_2")
horizontalLayout_17 = QtGui.QHBoxLayout(groupBox_2)
horizontalLayout_17.setObjectName("horizontalLayout_17")
pushButton_4 = QtGui.QPushButton(groupBox_2)
pushButton_4.setObjectName("pushButton_4")
horizontalLayout_17.addWidget(pushButton_4)
verticalLayout_8.addWidget(groupBox_2)
spacerItem6 = QtGui.QSpacerItem(20, 40, QtGui.QSizePolicy.Minimum, QtGui.QSizePolicy.Expanding)
verticalLayout_8.addItem(spacerItem6)
tabWidget.addTab(tab_3, (""))
'''-----------------------МЕНЮ СТАТУС БАР И ПРОЧЕЕ-------------------------------'''
verticalLayout.addWidget(tabWidget)
horizontalLayout.addLayout(verticalLayout)
LCSBSuite.setCentralWidget(centralwidget)
menubar = QtGui.QMenuBar(LCSBSuite)
menubar.setGeometry(QtCore.QRect(0, 0, 776, 26))
menubar.setObjectName("menubar")
menu = QtGui.QMenu(menubar)
menu.setObjectName("menu")
menu_2 = QtGui.QMenu(menubar)
menu_2.setObjectName("menu_2")
LCSBSuite.setMenuBar(menubar)
action = QtGui.QAction(LCSBSuite)
action.setObjectName("action")
action_2 = QtGui.QAction(LCSBSuite)
action_2.setObjectName("action_2")
menu.addAction(action)
menu_2.addAction(action_2)
menubar.addAction(menu_2.menuAction())
menubar.addAction(menu.menuAction())
statusBar = QtGui.QStatusBar(LCSBSuite)
statusBar.setObjectName(("statusBar"))
LCSBSuite.setStatusBar(statusBar)


LCSBSuite.setWindowTitle("LCSBSuite")
CSVgroupBox.setTitle("CSV Processing")
selcsvButton.setText(u"Выберите файлы .csv с \n данными по входным образцам")
resDSButton.setText(u"Выберите файл результата с \n Trinean DropSense")
wlDSButton.setText(u"Сформировать файл для \n Trinean DropSense")
startcell.setText(u"Стартовая ячейка:")
dataButton.setText(u"Перенести данные \n в исходный файл")
label_2.setText(u"Обьем образца до")
label_5.setText("ul")
seqRButton.setText("Name, Sequece")
mwRButton.setText("Name, MW, Ext.Coef")
label_3.setText(u"Концентрация до")
label_4.setText("pmol/ul")
Re_Button.setText(u"Сформировать Re файлы")
F_Button.setText(u"Сформировать F файлы")
txtgroupBox.setTitle("TXT Processing")
txtopenButton.setText(u"Выберите текстовый файл \n Worklist синтеза")
txtsaveButton.setText(u"Сохранить Worklist для DropSense \n в текстовом формате")
tabWidget.setTabText(tabWidget.indexOf(tab), ("DropSense"))
graphBox.setTitle(u"Графики")
showgrafButton.setText(u"Показать  графики")
csvBox.setTitle(u"CSV Статистика")
dirButton.setText(u"Выбрать директорию")
misButton.setText(u"Проверить на ошибки")
label.setText(u"Количество дней")
intervalButton.setText(u"Установить интервал")
intstatButton.setText(u"Рассчитать статистику \n интервала")
dirstatButton.setText(u"Рассчитать статистику \n директории")
excelBox.setTitle(u"Excel статистика")
excelopenButton.setText(u"Открыть файл")
excelButton2.setText(u"Рассчитать")
tabWidget.setTabText(tabWidget.indexOf(tab_2), (u"Статистика"))
groupBox.setTitle(u"Рассчёт реагентов на синтез")
loadWLButton.setText(u"Загрузить worklist \n синтеза")
label_6.setText(u"Масштаб \n синтеза")
label_7.setText(u"№ прибора")
pushButton_2.setText(u"Рассчитать \n реагенты")
groupBox_3.setTitle(u"Протоколы синтеза")
label_8.setText(u"Тип протокола")
pushButton_3.setText(u"Заполнить поля")
groupBox_2.setTitle(u"Утилита")
pushButton_4.setText(u"Press Button")
tabWidget.setTabText(tabWidget.indexOf(tab_3), u"Утилиты")
menu.setTitle(u"Помощь")
menu_2.setTitle(u"Файл")
action.setText(u"О программе")
action_2.setText(u"Выход")
LCSBSuite.show()
sys.exit(app.exec_())

